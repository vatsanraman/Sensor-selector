#! /usr/bin/python

from phil import *
from openpyxl.reader.excel import load_workbook
from alt_format_plate_wells_to_excel_sheet_mapping import plate_to_excel
import decimal
from TimeCourse import *

class Experiment:
        
    def __init__(self, name):
        self.name = name
        self.preextracted_file_exists = False
        print "Initializing ",self.name
        self.preextracted_file = '.'+self.name[:-5]+'.txt'
        if not os.path.isfile(self.preextracted_file):
            print "Pre extracted file does not exist"
            self.preextract_file()
        else:
            print "Read from pre-extracted file"
            self.__open_preextracted_file()
            self.preextracted_file_exists = True
    
    def __open_worksheet(self):
        wb = load_workbook(self.name)
        ws = wb.worksheets[0]#uses only first worksheet
        return ws
    
    def total_timepoints(self):
        return self.__open_worksheet().get_highest_row()

    def total_timecourses(self):
        return self.__open_worksheet().get_highest_column()
    
    def extract_timecourse(self,well):
        if self.preextracted_file_exists:
            return self.__preextracted_file_extract_timecourse(well)
            
        ws = self.__open_worksheet()
        timecourse = []
        timepoints = self.total_timepoints()
        for i in range(1, timepoints + 1):
            try:
                cell_address = '%s%s'%(plate_to_excel[well],i)
            except KeyError:
                print "Out of bound well address hit. Exiting.."
                exit()
            cell_value = ws.cell(cell_address).value
            d = decimal.Decimal(cell_value) 
            timecourse.append(float('%1.4f'%d))
        return timecourse

    def extract_all_timecourses(self):
        test_key_set = ['A1', 'A2', 'A3', 'A4']### REPLACE WITH FULL SET
        all_timecourses = {}
        for item in plate_to_excel:
            if not all_timecourses.has_key(item):
                all_timecourses[item] = TimeCourse(item, self.extract_timecourse(item))
        self.all_timecourses = all_timecourses

    def preextract_file(self):
        self.extract_all_timecourses()
        data = open(self.preextracted_file,'w')
        for keys in self.all_timecourses:
            data.write('%s '%keys)
            for item in self.all_timecourses[keys].data():
                data.write('%1.4f '%item)
            data.write('\n')

    def __open_preextracted_file(self):
        test_key_set = ['A1', 'A2', 'A3', 'A4']
        all_timecourses = {}
        data = open(self.preextracted_file, 'r')
        line = data.readline()
        while line:
            l = string.split(line)
            tmp = []
            tmp = [float(i) for i in l[1:]]
            if not all_timecourses.has_key(l[0]):
                all_timecourses[l[0]] = TimeCourse(l[0],tmp)
            line = data.readline()
        self.all_timecourses = all_timecourses
        
    
    def __preextracted_file_extract_timecourse(self, well):
        return self.all_timecourses[well].data()
    


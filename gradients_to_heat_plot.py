#! /usr/bin/python

from phil import *
from openpyxl.reader.excel import load_workbook
from alt_format_plate_wells_to_excel_sheet_mapping import plate_to_excel
import decimal
import stats
#minor changes

#imports for matplotlib
import numpy as np
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt



def Help():
    print 'Converts plate reader excel sheet data to heat plot\n'
    print 'Usage:\n'
    print '<input excel file, only OD data>'
    print '<selection gradient file, selection conc in each plate well>'
    print '<inducer gradient file, inducer conc in each plate well>'
    print '<number of time points to be considered>'
    print '<OD value computed as max or median>'
    print '<output heatmap file name, don\'t include extension>\n'
    exit()

if len(argv) < 2:
    Help()

inputfile = sys.argv[1]
selection_gradient = sys.argv[2]
inducer_gradient = sys.argv[3]
data_row = int(sys.argv[4])
characteristic_value_type = sys.argv[5]
output_file = sys.argv[6]


characteristic_value_of_time_course = {}#this could be max OD, median OD, slope etc

def parse_excel_sheet():
    wb = load_workbook(inputfile)
    ws = wb.worksheets[0]
    max_rows = ws.get_highest_row()
    max_columns = ws.get_highest_column()
    try:
        assert(data_row <= max_rows)
    except AssertionError:
        print "Total number of time points is lesser than your requested time point"
        exit()
    plate_rows = ['A','B','C','D','E','F','G','H']
    plate_data = {}
    for items in plate_rows:
        for i in range(1,13):
            plate_cell = '%s%s'%(items,i)
            if not plate_data.has_key(plate_cell):
                plate_data[plate_cell] = []
            for j in range(2, data_row + 1):
                cell_address = '%s%s'%(plate_to_excel[plate_cell], j)
                cell_value = ws.cell(cell_address).value
                try:
                    assert(cell_value)
                except AssertionError:
                    print "There is no data in cell %s"%cell_address
                    exit()
                d = decimal.Decimal(cell_value)
                plate_data[plate_cell].append(float('%1.5f'%d))

    return plate_data

#Overloaded function
def characteristic_value(plate_data):
    for keys in plate_data:
        if not characteristic_value_of_time_course.has_key(keys):
            if characteristic_value_type == 'median':
                characteristic_value_of_time_course[keys] = stats.lmedian(plate_data[keys])
            else:
                characteristic_value_of_time_course[keys] = max(plate_data[keys])
    for keys in characteristic_value_of_time_course:
        print keys, characteristic_value_of_time_course[keys]

#Overloaded function
def characteristic_value(OD_list):
    if characteristic_value_type == 'median':
        return float('%1.4f'%(stats.lmedian(OD_list)))
    elif characteristic_value_type == 'max':
        return float('%1.4f'%(max(OD_list)))
    else:
        print 'characteristic value type not defined, enter median or max'
        exit()
    


def combine_OD_selection_inducer_data(plate_data):
    OD_selection_inducer_data = {}
    
    for keys in plate_data:
        if not OD_selection_inducer_data.has_key(keys):
            OD_selection_inducer_data[keys] = []
        OD_selection_inducer_data[keys].append(characteristic_value(plate_data[keys]))

    selection_data = open(selection_gradient,'r')
    line = selection_data.readline()
    while line:
        l = string.split(line)
        if l[0] in OD_selection_inducer_data.keys():
            OD_selection_inducer_data[l[0]].insert(0,l[1])
        line = selection_data.readline()
    
    inducer_data = open(inducer_gradient,'r')
    line = inducer_data.readline()
    while line:
        l = string.split(line)
        if l[0] in OD_selection_inducer_data.keys():
            OD_selection_inducer_data[l[0]].insert(0,l[1])
        line = inducer_data.readline()
    
    return OD_selection_inducer_data

def normList(L, normalizeTo=1):
    vMax = max(L)
    return [ float('%1.4f'%(x/(vMax*1.0)*normalizeTo)) for x in L]


def generate_heat_plot(_OD_selection_inducer_data):
    selection = []
    inducer = []
    OD = []
    normalized_OD = []
    for keys in _OD_selection_inducer_data:
        inducer.append(float(_OD_selection_inducer_data[keys][0]))
        selection.append(float(_OD_selection_inducer_data[keys][1]))
        OD.append(float(_OD_selection_inducer_data[keys][2]))
    
    normalized_OD = normList(OD) #NOTE: we are plotting OD normalized to 1

    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.set_xscale('log')
    ax.set_yscale('log')
    ax.scatter(selection,inducer,c=normalized_OD,cmap = plt.jet(), s=670, lw = 0.1, marker = 's')
    bar = fig.colorbar(ax.collections[0])
    plt.show()
    fig.savefig('%s.png'%output_file)
    
    

_plate_data = {}
_OD_selection_inducer_data = {}
_plate_data = parse_excel_sheet()
_OD_selection_inducer_data = combine_OD_selection_inducer_data(_plate_data)
generate_heat_plot(_OD_selection_inducer_data)


#print _OD_selection_inducer_data
